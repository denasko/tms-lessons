import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active
sheet['A1'] = 'Name'
sheet['B1'] = 'Surname'
sheet['C1'] = 'Gender'
sheet['A2'] = 'Den'
sheet['B2'] = 'Aviz'
sheet['C2'] = 'M'

wb.save('file_09.xlsx')



wb1 = openpyxl.Workbook()

sheet = wb1.active
sheet['A1'] = 'Name'
sheet['B1'] = 'Surname'
sheet['C1'] = 'Old'
sheet['A2'] = input('Enter your name')
sheet['B2'] = input('Enter yout surname')
sheet['C2'] = input('Enter your years old')

wb1.save('file_10.xlsx')



wb3 = openpyxl.load_workbook('file_10.xlsx')
sheet = wb3.active
print(f'{sheet["A2"].value} {sheet["B2"].value} {sheet["C2"].value}')
